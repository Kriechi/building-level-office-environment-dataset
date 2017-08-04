#!/usr/bin/env python3

import os
import json
from openpyxl import load_workbook

APPLIANCE_LOG_FILENAME = os.path.join(os.environ['LOCAL_PATH_PREFIX'], 'Appliance Log.xlsx')

wb = load_workbook(filename=APPLIANCE_LOG_FILENAME, read_only=True)

appliance_log = dict()

for name in wb.get_sheet_names():
    ws = wb[name]

    log = dict()
    log['circuit_id'] = ws['A7'].value.split(' ')[-1]

    entries = []
    for r in range(10, ws.max_row, 3):
        entry = dict()
        entry_date = ws['A{}'.format(r)].value
        if not entry_date:
            break

        date = ws['A{}'.format(r)].value
        time = ws['A{}'.format(r + 1)].value
        entry['timestamp'] = '{}{}'.format(date.strftime('%Y-%m-%d'), time.strftime('T%H-%M-%S'))

        for i, c in enumerate('BCDEFG'):
            i += 1
            class_name = ws['{}{}'.format(c, r)].value
            appliance_name = ws['{}{}'.format(c, r + 1)].value
            power = ws['{}{}'.format(c, r + 2)].value

            if appliance_name == 'X':
                appliance_name = None

            if appliance_name == '✓':
                entry['socket_{}'.format(i)] = entries[-1]['socket_{}'.format(i)]
            else:
                entry['socket_{}'.format(i)] = {
                    'class_name': class_name,
                    'appliance_name': appliance_name,
                    'power': power,
                }

        entries.append(entry)

    log['entries'] = entries
    appliance_log[name] = log

print(json.dumps(appliance_log, sort_keys=True))
